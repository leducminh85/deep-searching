import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from googleapiclient.discovery import build
import concurrent.futures
import re
import os
import time
import json
import random
import threading
import yt_dlp
import ollama
from dotenv import load_dotenv

# Load môi trường từ .env
load_dotenv()

# --- CẤU HÌNH ---
FILE_PATH = "data.xlsx"
CHANNEL_LIST_SHEET = "Channel Lisitng" 
CACHE_SHEET = "NEW_CACHE_DATA_HIDDEN_"
YOUTUBE_MAX_WORKERS = 3 
AI_MAX_WORKERS = 2      # Số luồng AI (nên để thấp để tránh tràn VRAM)
SAVE_EVERY = 5
COOKIES_FILE = "cookies.txt"

API_KEY = os.getenv("YOUTUBE_API_KEY")

# Biến toàn cục cho AI
stop_flag = False
lock = threading.Lock()
save_lock = threading.Lock() 
ai_success_count = 0

def clean_for_excel(text):
    if not isinstance(text, str): return text
    text = ILLEGAL_CHARACTERS_RE.sub('', text)
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', text)
    return text[:32000]

def parse_duration(duration_str):
    hours = re.search(r'(\d+)H', duration_str)
    minutes = re.search(r'(\d+)M', duration_str)
    seconds = re.search(r'(\d+)S', duration_str)
    total_seconds = 0
    if hours: total_seconds += int(hours.group(1)) * 3600
    if minutes: total_seconds += int(minutes.group(1)) * 60
    if seconds: total_seconds += int(seconds.group(1))
    return total_seconds

# --- PHẦN 1: LẤY DỮ LIỆU TỪ YOUTUBE API ---

def get_channel_id_and_uploads_id(youtube, url):
    try:
        handle_match = re.search(r'(@[\w.-]+)', url)
        if handle_match:
            res = youtube.channels().list(part="contentDetails", forHandle=handle_match.group(1)).execute()
            if res.get('items'): return res['items'][0]['id'], res['items'][0]['contentDetails']['relatedPlaylists']['uploads']
        
        id_match = re.search(r'channel/(UC[\w-]{22})', url)
        if id_match:
            res = youtube.channels().list(part="contentDetails", id=id_match.group(1)).execute()
            if res.get('items'): return id_match.group(1), res['items'][0]['contentDetails']['relatedPlaylists']['uploads']

        user_match = re.search(r'user/([\w.-]+)', url)
        if user_match:
            res = youtube.channels().list(part="contentDetails", forUsername=user_match.group(1)).execute()
            if res.get('items'): return res['items'][0]['id'], res['items'][0]['contentDetails']['relatedPlaylists']['uploads']
    except Exception as e:
        print(f"❌ Lỗi lấy ID: {e}")
    return None, None

def get_all_videos_from_playlist(youtube, playlist_id, channel_name):
    videos = []
    next_page_token = None
    try:
        while True:
            res = youtube.playlistItems().list(part="snippet,contentDetails", playlistId=playlist_id, maxResults=50, pageToken=next_page_token).execute()
            for item in res.get('items', []):
                snippet = item['snippet']
                v_id = snippet['resourceId']['videoId']
                videos.append({
                    'id': v_id,
                    'Title': clean_for_excel(snippet['title']),
                    'URL': f"https://www.youtube.com/watch?v={v_id}",
                    'Thumbnail': snippet.get('thumbnails', {}).get('high', {}).get('url'),
                    'Date Published': snippet['publishedAt'],
                    'Channel Name': clean_for_excel(channel_name or snippet['channelTitle'])
                })
            next_page_token = res.get('nextPageToken')
            if not next_page_token: break
            
        filtered = []
        for i in range(0, len(videos), 50):
            batch = videos[i:i+50]
            v_res = youtube.videos().list(part="statistics,contentDetails", id=",".join([v['id'] for v in batch])).execute()
            stats_map = {item['id']: item for item in v_res.get('items', [])}
            for v in batch:
                data = stats_map.get(v['id'])
                if data and parse_duration(data['contentDetails'].get('duration', '')) > 180:
                    v['Views'] = int(data['statistics'].get('viewCount', 0))
                    filtered.append(v)
        return filtered
    except Exception as e:
        print(f"❌ Lỗi playlist {playlist_id}: {e}")
    return []

def process_channel_batch(channel):
    youtube = build("youtube", "v3", developerKey=API_KEY)
    ch_id, uploads_id = get_channel_id_and_uploads_id(youtube, channel['url'])
    return get_all_videos_from_playlist(youtube, uploads_id, channel['name']) if uploads_id else []

# --- PHẦN 2: PHÂN TÍCH AI (ANALYSIS) ---

def fetch_caption(video_url, row_idx):
    ydl_opts = {'skip_download': True, 'quiet': True, 'no_warnings': True}
    if os.path.exists(COOKIES_FILE): ydl_opts['cookiefile'] = COOKIES_FILE
    try:
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(video_url, download=False)
            subs = info.get('subtitles', {})
            auto = info.get('automatic_captions', {})
            sub_url = None
            for lang in [k for k in {**subs, **auto}.keys() if k.startswith('en')]:
                for f in {**subs, **auto}[lang]:
                    if f.get('ext') in ['json3', 'json']:
                        sub_url = f['url']; break
                if sub_url: break
            if not sub_url: return None
            sub_data = yt_dlp.YoutubeDL(ydl_opts).urlopen(sub_url).read().decode('utf-8')
            res = json.loads(sub_data)
            text = "".join([s['utf8'] for e in res.get('events',[]) if 'segs' in e for s in e['segs'] if 'utf8' in s])
            return ' '.join(text.split())[:30000]
    except Exception as e:
        if "429" in str(e): return "IP_BLOCKED"
        return None

def generate_summary(caption_text):
    try:
        response = ollama.chat(model='llama3.2:3b', messages=[{'role': 'user', 'content': f"Analyze this YouTube transcript (Main Story, Characters, Location, Conflict Type, Keywords):\n\n{caption_text}"}], options={'num_ctx': 8192})
        return response['message']['content'].strip()
    except: return "ERROR_AI"

def ai_process_row(row_idx, url, existing_caption, existing_summary, ws, caption_col, summary_col):
    global stop_flag, ai_success_count
    if stop_flag: return
    time.sleep(random.uniform(2, 5))
    cap = existing_caption
    if not cap or cap == "" or "ERROR" in str(cap):
        cap = fetch_caption(url, row_idx)
        if cap == "IP_BLOCKED": stop_flag = True; return
    
    summ = existing_summary
    if cap and cap != "#" and (not summ or summ == "" or "ERROR" in str(summ)):
        summ = generate_summary(cap)
    
    with lock:
        ws.cell(row=row_idx, column=caption_col).value = clean_for_excel(cap or "#")
        ws.cell(row=row_idx, column=summary_col).value = clean_for_excel(summ or "#")
        ai_success_count += 1

# --- MAIN WORKFLOW ---

def main():
    if not API_KEY: return print("❌ API KEY missing")
    print(f"📖 Reading {FILE_PATH}...")
    wb = openpyxl.load_workbook(FILE_PATH)
    
    # 1. YouTube Fetching
    channel_ws = wb[CHANNEL_LIST_SHEET]
    channels = [{'url': row[0].value, 'name': row[1].value} for row in channel_ws.iter_rows(min_row=2) if row[0].value]
    
    all_videos = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=YOUTUBE_MAX_WORKERS) as executor:
        futures = [executor.submit(process_channel_batch, ch) for ch in channels]
        for f in concurrent.futures.as_completed(futures): all_videos.extend(f.result() or [])

    cache_ws = wb[CACHE_SHEET]
    headers = [str(cell.value) for cell in cache_ws[1]]
    cols = {h: i+1 for i, h in enumerate(headers)}
    
    existing_urls = {str(cache_ws.cell(row=r, column=cols['URL']).value): r for r in range(2, cache_ws.max_row + 1)}

    for vid in all_videos:
        r = existing_urls.get(vid['URL'], cache_ws.max_row + 1)
        for k, v in vid.items(): 
            if k in cols: cache_ws.cell(row=r, column=cols[k]).value = v
        if vid['URL'] not in existing_urls: 
            cache_ws.cell(row=r, column=cols['URL']).value = vid['URL']
            existing_urls[vid['URL']] = r
    
    wb.save(FILE_PATH)
    print(f"✅ YouTube sync done. Saved to Excel.")

    # 2. AI Analysis
    print(f"🤖 Starting AI Analysis...")
    tasks = []
    for r in range(2, cache_ws.max_row + 1):
        url = cache_ws.cell(row=r, column=cols['URL']).value
        cap = cache_ws.cell(row=r, column=cols['Caption']).value
        summ = cache_ws.cell(row=r, column=cols['Summary']).value
        if url and (not cap or not summ or "ERROR" in str(cap) or "ERROR" in str(summ)):
            tasks.append((r, url, cap, summ))

    with concurrent.futures.ThreadPoolExecutor(max_workers=AI_MAX_WORKERS) as executor:
        for r, url, cap, summ in tasks:
            executor.submit(ai_process_row, r, url, cap, summ, cache_ws, cols['Caption'], cols['Summary'])
            if ai_success_count > 0 and ai_success_count % SAVE_EVERY == 0:
                with save_lock: wb.save(FILE_PATH)
    
    wb.save(FILE_PATH)
    print(f"🎉 All processes completed!")

if __name__ == "__main__":
    main()
