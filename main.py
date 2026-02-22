import time
import random
import threading
import concurrent.futures
import yt_dlp
import requests
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import re
import ollama

# --- CẤU HÌNH CƠ BẢN ---
FILE_PATH = "data.xlsx"   
SHEET_NAME = "NEW_CACHE_DATA_HIDDEN_"
MAX_WORKERS = 1       # Bắt buộc để 1 luồng khi chạy AI Local để tránh treo máy
SAVE_EVERY = 2        
MAX_TEST_VIDEOS = 1  # Số lượng video muốn chạy thử. Đổi thành số lớn hơn nếu muốn chạy thật.

# Biến toàn cục
stop_flag = False
lock = threading.Lock()
save_lock = threading.Lock() 
success_count = 0
wb = None
ws = None

def clean_for_excel(text):
    """Tẩy rửa triệt để các ký tự rác làm hỏng cấu trúc XML của Excel"""
    if not isinstance(text, str):
        return text
        
    # 1. Dùng regex chuẩn của openpyxl để lọc toàn bộ ký tự cấm làm corrupt file
    text = ILLEGAL_CHARACTERS_RE.sub('', text)
    
    # 2. Lọc thêm một lớp dự phòng các ký tự điều khiển
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', text)
    
    # 3. Đảm bảo an toàn tuyệt đối độ dài (Excel tối đa khoảng 32,767 ký tự/ô)
    return text[:32000]

def fetch_caption(video_url, row_idx):
    """Lấy caption dùng yt-dlp, xử lý triệt để lỗi Format Video"""
    ydl_opts = {
            'skip_download': True,
            'quiet': True,
            'no_warnings': True,
            'ignore_no_formats_error': True,
        }
    
    try:
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(video_url, download=False)
            
        if not info:
            print(f"[Dòng {row_idx}] ❌ Lỗi: Không thể truy cập dữ liệu video.")
            return None
            
        subs = info.get('subtitles', {})
        auto_subs = info.get('automatic_captions', {})
        
        sub_url = None
        
        # 1. Tìm phụ đề thủ công
        if 'en' in subs:
            for f in subs['en']:
                if f.get('ext') == 'json3':
                    sub_url = f['url']
                    break
                    
        # 2. Tìm phụ đề tự động (auto-generated)
        if not sub_url:
            for lang in auto_subs.keys():
                if lang.startswith('en'):  
                    for f in auto_subs[lang]:
                        if f.get('ext') == 'json3':
                            sub_url = f['url']
                            break
                    if sub_url: break

        if not sub_url:
            print(f"[Dòng {row_idx}] ⚠️ Lỗi: Video KHÔNG CÓ phụ đề tiếng Anh.")
            return None
            
        # 3. Tải và xử lý file phụ đề
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
            }
            response = requests.get(sub_url, headers=headers)
            
            if response.status_code != 200:
                print(f"[Dòng {row_idx}] ❌ Lỗi: YouTube từ chối cho tải file phụ đề (Mã lỗi {response.status_code}).")
                if response.status_code == 429:
                    return "IP_BLOCKED"
                return None
                
            res = response.json()
            
        except requests.exceptions.JSONDecodeError:
            print(f"[Dòng {row_idx}] ❌ Lỗi cấu trúc: YouTube trả về dữ liệu rỗng hoặc sai định dạng.")
            return None
            
        text_chunks = []
        for event in res.get('events', []):
            if 'segs' in event:
                for seg in event['segs']:
                    if 'utf8' in seg:
                        text_chunks.append(seg['utf8'])
                        
        full_text = "".join(text_chunks).replace('\n', ' ')
        full_text = ' '.join(full_text.split()) 
        
        return full_text[:30000] # Cắt bớt nếu quá dài để tránh tràn RAM khi đưa vào AI
        
    except Exception as e:
        error_msg = str(e).lower()
        if "unavailable" in error_msg or "private" in error_msg or "removed" in error_msg:
            print(f"[Dòng {row_idx}] ❌ Lỗi: Video không tồn tại, bị xóa hoặc đã chuyển sang riêng tư.")
        elif "too many requests" in error_msg or "http error 429" in error_msg:
            return "IP_BLOCKED"
        elif "sign in" in error_msg or "age" in error_msg or "bot" in error_msg:
            print(f"[Dòng {row_idx}] ❌ Lỗi: Video giới hạn độ tuổi, Cookie hết hạn hoặc bắt xác minh danh tính.")
        else:
            print(f"[Dòng {row_idx}] ❌ Lỗi yt-dlp: {str(e).splitlines()[0][:80]}")
        return None

def generate_summary(caption_text, row_idx):
    """Sử dụng Ollama (Local AI) để đọc caption và trích xuất cấu trúc câu chuyện"""
    if not caption_text or caption_text == "#" or len(str(caption_text).strip()) < 50:
        return "#"
        
    prompt = f"""
    You are a video data analyst. Read the YouTube caption below and summarize its content for in-depth search.
    Please return results strictly following this format (write concisely, directly in Vietnamese, except for NICHE KEYWORDS which remain in English):
    1. MAIN STORY: (Briefly describe what happened in 2-3 sentences).
    2. CHARACTERS & RELATIONSHIPS: (Who was involved? What was their relationship?).
    3. LOCATION / CONTEXT: (Where did the event take place?).
    4. TYPE OF CONFLICT: (Conflict over property, verbal, physical, legal, etc.).
    5. NICHE KEYWORDS (TAGS): (List 5-7) English keywords that accurately describe the video's niche. For example: neighbor dispute, crazy ex, public freakout, karen, road rage.
    Caption:
    {caption_text[:30000]}
    """
    
    try:
        # Gọi trực tiếp model Llama 3.1 đã tải trên máy
        response = ollama.chat(model='llama3.1', messages=[
            {
                'role': 'user',
                'content': prompt,
            },
        ])
        return response['message']['content'].strip()
    except Exception as e:
        print(f"[Dòng {row_idx}] ❌ Lỗi Local AI: Có thể Ollama chưa chạy hoặc model quá nặng. Chi tiết: {e}")
        return "ERROR_AI"

def process_row(row_idx, url, existing_caption, existing_summary):
    """Xử lý từng dòng: Lấy caption (nếu chưa có) -> Tạo summary (nếu chưa có)"""
    global stop_flag, success_count
    
    if stop_flag:
        return row_idx, "ABORTED", "ABORTED"
    
    # Nghỉ dài hơn chút để chống lỗi 429 từ YouTube
    time.sleep(random.uniform(3, 7))
    
    print(f"\n▶ Đang xử lý dòng {row_idx}...")
    
    final_caption = existing_caption
    final_summary = existing_summary
    
    # 1. Xử lý Caption
    if existing_caption is None or str(existing_caption).strip() == "" or "ERROR" in str(existing_caption):
        print(f"  [Dòng {row_idx}] Đang tải phụ đề...")
        final_caption = fetch_caption(url, row_idx)
        
        if final_caption == "IP_BLOCKED":
            print(f"  [CẢNH BÁO TỚI HẠN] ⛔ IP bị YouTube chặn ở dòng {row_idx}!")
            stop_flag = True 
            return row_idx, "ABORTED", "ABORTED"
            
        if not final_caption:
            final_caption = "#"

    # 2. Xử lý Summary (Chỉ chạy khi đã có caption và chưa có summary hợp lệ)
    if final_caption and final_caption != "#" and final_caption != "ERROR":
        if existing_summary is None or str(existing_summary).strip() == "" or "ERROR" in str(existing_summary):
            print(f"  [Dòng {row_idx}] 🤖 AI đang đọc và tóm tắt nội dung (có thể mất vài chục giây)...")
            final_summary = generate_summary(final_caption, row_idx)
    else:
        final_summary = "#" # Nếu video lỗi không có caption thì summary cũng bỏ qua

    with lock:
        success_count += 1
        
    return row_idx, clean_for_excel(final_caption), clean_for_excel(final_summary)

def main():
    global success_count, stop_flag, wb, ws  
    
    print(f"Đang đọc dữ liệu từ {FILE_PATH}...")
    try:
        wb = openpyxl.load_workbook(FILE_PATH)
        ws = wb[SHEET_NAME]
    except Exception as e:
        print(f"Lỗi đọc file: {e}. Vui lòng ĐÓNG KÍN file Excel trước khi chạy.")
        return

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    
    try:
        url_col_idx = headers.index('URL') + 1
    except ValueError:
        print("Lỗi: Không tìm thấy cột 'URL' ở dòng đầu tiên của Excel!")
        return
        
    # Tạo cột Caption nếu chưa có
# Giả sử headers là danh sách tiêu đề hiện tại: headers = [cell.value for cell in ws[1]]

    # 1. Xử lý cột Caption (Cột E - Vị trí 5)
    if 'Caption' in headers:
        caption_col_idx = headers.index('Caption') + 1
    else:
        caption_col_idx = 5
        ws.insert_cols(caption_col_idx)  # Chèn một cột trống tại vị trí E
        ws.cell(row=1, column=caption_col_idx).value = 'Caption'
        # Cập nhật lại danh sách headers sau khi chèn để tính toán chính xác cho cột tiếp theo
        headers = [cell.value for cell in ws[1]]

    # 2. Xử lý cột Summary (Cột F - Vị trí 6)
    if 'Summary' in headers:
        summary_col_idx = headers.index('Summary') + 1
    else:
        summary_col_idx = 6
        ws.insert_cols(summary_col_idx)  # Chèn một cột trống tại vị trí F
        ws.cell(row=1, column=summary_col_idx).value = 'Summary'
        headers = [cell.value for cell in ws[1]]

    wb.save(FILE_PATH)
    print(f"Đã đảm bảo các cột Caption (E) và Summary (F) có trong file Excel.")

    tasks = []
    for row_idx in range(2, ws.max_row + 1):
        url = ws.cell(row=row_idx, column=url_col_idx).value
        caption = ws.cell(row=row_idx, column=caption_col_idx).value
        summary = ws.cell(row=row_idx, column=summary_col_idx).value
        
        is_valid_url = isinstance(url, str) and 'youtube' in url.lower()
        
        # Chỉ lấy những dòng cần cập nhật Caption hoặc Summary
        needs_caption = caption is None or str(caption).strip() == "" or "ERROR" in str(caption)
        needs_summary = (summary is None or str(summary).strip() == "" or "ERROR" in str(summary)) and caption != "#"
        
        if is_valid_url and (needs_caption or needs_summary):
            tasks.append((row_idx, url, caption, summary))
            if len(tasks) >= MAX_TEST_VIDEOS:
                break
                
    print(f"Tìm thấy {len(tasks)} video cần xử lý.")
    
    if not tasks:
        print("Mọi video đều đã có đầy đủ dữ liệu! Hoàn thành.")
        return

    print(f"Bắt đầu chạy với {MAX_WORKERS} luồng...\n")
    
    try:
        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            future_to_index = {
                executor.submit(process_row, r_idx, url, cap, sum_): r_idx 
                for r_idx, url, cap, sum_ in tasks
            }
            
            for future in concurrent.futures.as_completed(future_to_index):
                r_idx, cap_result, sum_result = future.result()
                
                with save_lock:
                    if cap_result == "ABORTED":
                        continue
                        
                    # Cập nhật dữ liệu vào Excel
                    if cap_result:
                        ws.cell(row=r_idx, column=caption_col_idx).value = cap_result
                    if sum_result:
                        ws.cell(row=r_idx, column=summary_col_idx).value = sum_result
                        
                    if success_count > 0 and success_count % SAVE_EVERY == 0:
                        wb.save(FILE_PATH)
                        print(f" -> Đã lưu an toàn (Tổng: {success_count} dòng hoàn tất)!")

    except KeyboardInterrupt:
        print("\n[DỪNG KHẨN CẤP] Người dùng đã yêu cầu dừng chương trình!")
        stop_flag = True

    finally:
        print("\nĐang lưu lại toàn bộ dữ liệu lần cuối...")
        with save_lock:
            wb.save(FILE_PATH)
        print("Hoàn tất quy trình!")

if __name__ == "__main__":
    main()