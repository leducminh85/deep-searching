import openpyxl
import os

FILE_PATH = "data.xlsx"
SHEET_NAME = "NEW_CACHE_DATA_HIDDEN_"

def cleanup():
    if not os.path.exists(FILE_PATH):
        print(f"❌ Không tìm thấy file {FILE_PATH}")
        return

    print(f"📖 Đang mở file {FILE_PATH}...")
    try:
        wb = openpyxl.load_workbook(FILE_PATH)
        ws = wb[SHEET_NAME]
    except Exception as e:
        print(f"❌ Lỗi: {e}. Hãy đảm bảo bạn đã đóng file Excel.")
        return

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    
    try:
        caption_col = headers.index('Caption') + 1
        summary_col = headers.index('Summary') + 1
    except ValueError:
        print("❌ Không tìm thấy cột 'Caption' hoặc 'Summary'.")
        return

    print("🧹 Bắt đầu quét và xóa dấu #...")
    count = 0
    
    for row in range(2, ws.max_row + 1):
        caption_cell = ws.cell(row=row, column=caption_col)
        summary_cell = ws.cell(row=row, column=summary_col)

        # Kiểm tra và xóa dấu # ở cột Caption
        if str(caption_cell.value).strip() == "#":
            caption_cell.value = ""
            count += 1

        # Kiểm tra và xóa dấu # ở cột Summary
        if str(summary_cell.value).strip() == "#":
            summary_cell.value = ""
            count += 1

    if count > 0:
        print(f"✅ Đã dọn dẹp xong {count} ô chứa dấu #.")
        print("💾 Đang lưu lại script. Vui lòng chờ...")
        wb.save(FILE_PATH)
        print("🎉 Hoàn tất!")
    else:
        print("✨ Không tìm thấy ô nào chỉ chứa dấu #. File đã sạch!")

if __name__ == "__main__":
    cleanup()
